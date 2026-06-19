#!/usr/bin/env python3
"""
build_playbook.py
=================
Turn the flat backtest results (comparative_analysis_results.csv) into a
presentation-ready Excel workbook, strategy_playbook.xlsx, with four sheets:

  1. Recommendations  - the rank-1 strategy per market condition x portfolio
  2. Playbook         - every TLH-on strategy, composite-ranked within each cell
  3. TLH Value-Add    - delta metrics for each strategy vs its paired No-TLH run
  4. Raw Data         - all 384 runs (TLH-on + No-TLH) for analysis

Composite score weights: Tax Alpha 30%, Sharpe 25%, CAGR 20%, Max Drawdown 15%,
Information Ratio 10% (each rank-normalised within Portfolio x Market Condition).

Run after run_backtest.py:  python Backtest/build_playbook.py
"""
from __future__ import annotations
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
PERIOD_ORDER = {'Bear Market': 0, 'Baseline Market': 1, 'Bull Market': 2,
                'Past 5 Years': 3, 'Past 10 Years': 4, 'Past 20 Years': 5}


def build_playbook_sheet(export_df: pd.DataFrame) -> pd.DataFrame:
    pb = export_df[export_df['TLH Status'] == 'On (10%)'].copy()
    g = pb.groupby(['Portfolio', 'Market Condition'])
    pb['_rank_sharpe'] = g['Sharpe Ratio'].rank(ascending=True)
    pb['_rank_cagr'] = g['CAGR'].rank(ascending=True)
    pb['_rank_dd'] = g['Max Drawdown'].rank(ascending=False)
    pb['_ir_filled'] = pb['Information Ratio'].fillna(0)
    pb['_rank_ir'] = pb.groupby(['Portfolio', 'Market Condition'])['_ir_filled'].rank(ascending=True)
    pb['_ta_filled'] = pb['Tax Alpha 2'].fillna(0)
    pb['_rank_tax_alpha'] = pb.groupby(['Portfolio', 'Market Condition'])['_ta_filled'].rank(ascending=True)
    pb['Composite Score'] = (
        0.30 * pb['_rank_tax_alpha'] + 0.25 * pb['_rank_sharpe'] +
        0.20 * pb['_rank_cagr'] + 0.15 * pb['_rank_dd'] + 0.10 * pb['_rank_ir']
    ).round(3)
    pb['Rank'] = (pb.groupby(['Portfolio', 'Market Condition'])['Composite Score']
                  .rank(ascending=False, method='min').astype(int))
    pb['_sort'] = pb['Market Condition'].map(PERIOD_ORDER).fillna(9)
    pb = pb.sort_values(['Portfolio', '_sort', 'Rank'])
    pb = pb.drop(columns=[c for c in pb.columns if c.startswith('_')])

    cols = ['Portfolio', 'Market Condition', 'Rank', 'Rebal Type', 'Rebal Value',
            'Strategy Label', 'CAGR', 'Sharpe Ratio', 'Max Drawdown',
            'Realized Losses (All)', 'TLH Event Count', 'Tax Alpha 2',
            'Tracking Error (Ann)', 'Information Ratio', 'Composite Score']
    out = pb[[c for c in cols if c in pb.columns]].reset_index(drop=True)
    for c in ['CAGR', 'Sharpe Ratio', 'Max Drawdown', 'Tracking Error (Ann)', 'Information Ratio']:
        if c in out.columns:
            out[c] = out[c].round(4)
    for c in ['Realized Losses (All)', 'Tax Alpha 2']:
        if c in out.columns:
            out[c] = out[c].round(2)
    return out


def build_value_add_sheet(export_df: pd.DataFrame) -> pd.DataFrame:
    on = export_df[export_df['TLH Status'] == 'On (10%)'].copy()
    off = export_df[export_df['TLH Status'] == 'Off'].copy()
    keys = ['Portfolio', 'Market Condition', 'Rebal Type', 'Rebal Value']
    va = on.merge(off[keys + ['CAGR', 'Sharpe Ratio', 'Max Drawdown', 'Final NAV']],
                  on=keys, suffixes=('', '_bench'))
    va['Delta CAGR'] = (va['CAGR'] - va['CAGR_bench']).round(4)
    va['Delta Sharpe'] = (va['Sharpe Ratio'] - va['Sharpe Ratio_bench']).round(4)
    va['Delta Max Drawdown'] = (va['Max Drawdown'] - va['Max Drawdown_bench']).round(4)
    va['Delta Final NAV'] = (va['Final NAV'] - va['Final NAV_bench']).round(2)
    va['_sort'] = va['Market Condition'].map(PERIOD_ORDER).fillna(9)
    va = va.sort_values(['Portfolio', '_sort', 'Rebal Type', 'Rebal Value']).drop(columns=['_sort'])
    cols = ['Portfolio', 'Market Condition', 'Rebal Type', 'Rebal Value', 'Strategy Label',
            'Delta CAGR', 'Delta Sharpe', 'Delta Max Drawdown', 'Delta Final NAV',
            'Realized Losses (All)', 'TLH Event Count', 'Tax Alpha 2',
            'Tracking Error (Ann)', 'Information Ratio']
    return va[[c for c in cols if c in va.columns]].reset_index(drop=True)


def build_decomposition_sheet(export_df: pd.DataFrame) -> pd.DataFrame:
    """Split each pair's net value-add into tax saved, replacement tracking, and cost."""
    on = export_df[export_df['TLH Status'] == 'On (10%)'].copy()
    off = export_df[export_df['TLH Status'] == 'Off'].copy()
    keys = ['Portfolio', 'Market Condition', 'Rebal Type', 'Rebal Value']
    d = on.merge(off[keys + ['Final NAV', 'Tax Paid', 'Execution Costs']],
                 on=keys, suffixes=('', '_off'))
    d['Net Value-Add'] = (d['Final NAV'] - d['Final NAV_off']).round(2)
    d['Tax Saved'] = (d['Tax Paid_off'] - d['Tax Paid']).round(2)
    d['Extra Cost'] = (d['Execution Costs_off'] - d['Execution Costs']).round(2)
    d['Replacement Tracking'] = (d['Net Value-Add'] - d['Tax Saved'] - d['Extra Cost']).round(2)
    d['_sort'] = d['Market Condition'].map(PERIOD_ORDER).fillna(9)
    d = d.sort_values(['Portfolio', '_sort', 'Rebal Type', 'Rebal Value']).drop(columns=['_sort'])
    cols = ['Portfolio', 'Market Condition', 'Strategy Label', 'Tax Saved',
            'Replacement Tracking', 'Extra Cost', 'Net Value-Add', 'TLH Event Count']
    return d[[c for c in cols if c in d.columns]].reset_index(drop=True)


def style_sheet(ws, df, header_hex, freeze_cols=2):
    header_fill = PatternFill('solid', fgColor=header_hex)
    header_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill = PatternFill('solid', fgColor='EEF2F7')
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal='center')
    for col_idx, col_name in enumerate(df.columns, 1):
        vals = [str(col_name)] + [str(v) for v in df.iloc[:, col_idx - 1]]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(v) for v in vals) + 2, 30)
    if freeze_cols:
        ws.freeze_panes = ws.cell(row=2, column=freeze_cols + 1)


def main():
    export_df = pd.read_csv(HERE / 'comparative_analysis_results.csv')

    pb_out = build_playbook_sheet(export_df)
    va_out = build_value_add_sheet(export_df)
    rec_cols = ['Portfolio', 'Market Condition', 'Strategy Label', 'Rebal Type', 'Rebal Value',
                'CAGR', 'Sharpe Ratio', 'Max Drawdown', 'Realized Losses (All)',
                'Tax Alpha 2', 'Composite Score']
    rec_out = pb_out[pb_out['Rank'] == 1][[c for c in rec_cols if c in pb_out.columns]].reset_index(drop=True)

    raw_out = export_df.copy()
    raw_out['_sort'] = raw_out['Market Condition'].map(PERIOD_ORDER).fillna(9)
    raw_out = (raw_out.sort_values(['Portfolio', '_sort', 'Rebal Type', 'Rebal Value', 'TLH Status'])
               .drop(columns=['_sort']).reset_index(drop=True))

    decomp_out = build_decomposition_sheet(export_df)

    colors = {'Recommendations': '1F3864', 'Playbook': '2E4057',
              'TLH Value-Add': '274060', 'TLH Decomposition': '1B4D3E', 'Raw Data': '3D3D3D'}
    data = {'Recommendations': rec_out, 'Playbook': pb_out, 'TLH Value-Add': va_out,
            'TLH Decomposition': decomp_out, 'Raw Data': raw_out}

    out_path = HERE / 'strategy_playbook.xlsx'
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for name, df in data.items():
            df.to_excel(writer, sheet_name=name, index=False)
        for ws in writer.book.worksheets:
            style_sheet(ws, data[ws.title], colors[ws.title])

    print(f'Wrote {out_path}')
    for name, df in data.items():
        print(f'  {name:16}: {len(df)} rows')


if __name__ == '__main__':
    main()
